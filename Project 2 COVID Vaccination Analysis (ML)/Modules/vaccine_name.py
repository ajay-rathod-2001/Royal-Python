# Vaccine name--> if day gap ==28, then COVAXIN else if day gap==84, then COVISHIELD
import openpyxl
from second_vaccination import find_path

excel_filename ="Vaccination_data.xlsx"

def get_day_gap():
    sfl = find_path()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    day_difference =[]
    
    for i in range(2,sheet.max_row+1):
        v =sheet.cell(row=i, column=4).value
        day_difference.append(v)
    return day_difference
    
def vaccine_naming():
    sfl = find_path()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    day_gapping =get_day_gap()
    l =len(day_gapping); start_row =2
    
    for i in range(0,l):
        if day_gapping[i]==84:
            sheet.cell(row=start_row+i,column=11, value ='COVISHIELD')
        else:
            sheet.cell(row=start_row+i,column=11, value ='COVAXIN')
    wb.save(sfl)
    
print("Appending file with confirmation for Vaccine Name.Please Wait!")
vaccine_naming()
print("Appended Successufully!!")