# 2nd vaccination - if 1st not vaccinated then no 2nd vaccination, or if vaccination  is confirmed then only confirm for 2nd vaccination

import openpyxl
from file_search import serach_file_location,serach_file_location_infolder

excel_filename ="Vaccination_data.xlsx"

def find_path():
    #  if Found in same folder
    res_in_folder,loc_in_folder = serach_file_location_infolder(excel_filename)
    
    if res_in_folder:
        source_file_location = loc_in_folder[0]
    else:
        # if not not found in the current directory,search else where
        res_elsewhere,loc_elsewhere = serach_file_location(excel_filename)
        
        if res_elsewhere:
            source_file_location = loc_elsewhere[0] # take the first location if multiple files are found (unlikely)
        else:
            print("Can't Detect the Location  of the file .")
            exit()
    return  source_file_location

def get_1st_confirm():
    sfl = find_path()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    first_vaccination_confirm =[]
    
    for i in range(2,sheet.max_row+1):
        v = sheet.cell(row = i, column= 8).value
        first_vaccination_confirm.append(v)
    return first_vaccination_confirm
    
def write_2nd_confirmation():
    yn1=get_1st_confirm()
    sfl = find_path()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    l =len(yn1); start_row =2
    
    for i in range(0,l):
        if yn1[i]=="Yes":
            sheet.cell(row=start_row+i,column=9, value ='Yes')
        else:
            sheet.cell(row=start_row+i,column=9, value ='No')
    wb.save(sfl)

if __name__=="__main__":         

    print("Appending file with confirmation for 2nd dose.Please wait!")
    # sfl=find_path()
    # print(sfl)
    # yn1=get_1st_confirm()
    # print(v)
    write_2nd_confirmation()
    print("Appeneded Successfully!!")
