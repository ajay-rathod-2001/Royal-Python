# Find the Booster Vaccination  Date

import openpyxl
from datetime import datetime
from dateutil.relativedelta import relativedelta
from file_search import serach_file_location,serach_file_location_infolder

excel_filename ="Vaccination_data.xlsx"

def find_path():
    #  if Found in same folder
    res_in_folder,loc_in_folder = serach_file_location_infolder(excel_filename)
    
    if res_in_folder:
        source_file_location = loc_in_folder[0]
    else:
        # if not not found in the current directory,search else where
        res_elsewhere,loc_elsewhere = serach_file_location(excel_filename)
        
        if res_elsewhere:
            source_file_location = loc_elsewhere[0] # take the first location if multiple files are found (unlikely)
        else:
            print("Can't Detect the Location  of the file .")
            exit()
    return  source_file_location

def extract_month_gap():
    sfl=find_path()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    month_gap = []
    
    for i in range (2,sheet.max_row+1):
        v = sheet.cell(row=i,column=6).value
        
        if v is not None:
            month_gap.append(v)
    month_gap = list(set(month_gap))
    if month_gap:
        return month_gap[0]
    else:
        return 6
def extract_2ndDose_Date():
    sfl=find_path()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    second_dose_date = []
    
    for i in range (2,sheet.max_row+1):
        v = sheet.cell(row=i,column=5).value
        
        if v is not None:
            second_dose_date.append(v)
    return second_dose_date

def boosterDoseDate():
   sfl=find_path()
   wb =openpyxl.load_workbook(sfl)
   sheet=wb.active
   sdd = extract_2ndDose_Date()
   six_mg = extract_month_gap()
   start_row = 2
   
   def get_sixMD(dt,six_mg):
       input_date=datetime.strptime(dt,"%d-%m-%Y")
       next_date = (input_date+relativedelta(months=six_mg)).strftime("%d-%m-%Y")
       return next_date
   for i,_ in enumerate(sdd):
    next_date=get_sixMD(sdd[i],six_mg)
    sheet.cell(row=start_row+i,column=7, value =next_date)
   wb.save(sfl)       
        
   


print("Appending file with Booster vaccination .Please waite.....")
# print(extract_month_gap())
# sdd =extract_2ndDose_Date()
# print(sdd)
boosterDoseDate()
print("File appended successfully!!")

# sfl=find_path()
# print(sfl)
            
    