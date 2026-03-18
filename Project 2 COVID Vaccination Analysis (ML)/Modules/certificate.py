# Final  column to be filled with vaccination Certificated recieved or not . Recieved = Yes if all the Doses taken
# Recieved =  No , if any dose is left

import openpyxl
from booster_vaccination import get_2nd_confirm
from second_vaccination import get_1st_confirm,find_path

excel_filename ="Vaccination_data.xlsx"

def get_booster_confirmation():
    sfl = find_path()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    booster_yn =[]
    
    for i in range(2,sheet.max_row+1):
        v = sheet.cell(row = i, column= 10).value
        booster_yn.append(v)
    return booster_yn

def certificate():
    sfl = find_path()
    yn1=get_1st_confirm()
    yn2=get_2nd_confirm()
    booster_yn =get_booster_confirmation()
    wb =openpyxl.load_workbook(sfl)
    sheet=wb.active
    l =len(yn1); start_row=2
    
    for i in range(0,l):
        if yn1[i]=="Yes" and yn2[i]=="Yes" and booster_yn[i]=="Yes" :
            sheet.cell(row=start_row+i,column=13, value ='Verified')
        else:
            sheet.cell(row=start_row+i,column=13, value ='Unverified')
    wb.save(sfl)
       
print("Appending file with Vaccination Certificate.Please wait!")
# v =get_booster_confirmation()
# print(v)
certificate()  
print("Appeneded Successfully!!")